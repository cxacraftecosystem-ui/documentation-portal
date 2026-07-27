"""The .xlsx report must never again open with Excel's "we found a problem" repair prompt.

An .xlsx is a zip of XML parts, and Excel offers to recover the file the moment one of them is
not well-formed or does not say what the schema expects. Eyeballing the download cannot catch
that: the workbook still opens, it just opens *repaired*, usually with the transcripts gone. So
this builds a workbook out of the worst input the live repository can realistically produce —
NUL and control bytes, a lone surrogate, Unicode noncharacters, a transcript past Excel's cell
ceiling, Devanagari, an emoji, sheet names carrying the characters Excel forbids in a tab title,
a field note that opens with "=", an empty rich-text run, a run with a broken font — and then
unpacks the archive and parses every part.

If a future change lets one of those through again, this fails here rather than on a craft
scholar's laptop with no way to tell what went wrong.
"""

import re
import posixpath
import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from app.services.transcript_format import transcript_cell
from app.services.xlsx_report import build_report_workbook

MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
CONTENT_TYPES = "{http://schemas.openxmlformats.org/package/2006/content-types}"
RELS = "{http://schemas.openxmlformats.org/package/2006/relationships}"

MAX_CELL_CHARS = 32767
TITLE_BAD = set(r"[]:*?/\\")

# Every codepoint XML 1.0 refuses inside element content. Expat — and Excel — reject a part
# containing any of these, whether it arrives raw or as a numeric character reference.
_ILLEGAL = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f"
    f"{chr(0xD800)}-{chr(0xDFFF)}"
    f"{chr(0xFDD0)}-{chr(0xFDEF)}"
    + "".join(chr(p * 0x10000 + o) for p in range(17) for o in (0xFFFE, 0xFFFF))
    + "]"
)
# "&#7;" / "&#55357;" — how ElementTree emits a character it cannot encode. No XML parser accepts
# a reference to a codepoint that is illegal in the first place, so the part stops parsing.
_CHAR_REF = re.compile(r"&#(\d+);")


# --------------------------------------------------------------------------------------------
# The input: what the field actually sends us.
# --------------------------------------------------------------------------------------------

# Long enough that the cell ceiling has to be enforced across rich-text runs, not per run.
LONG_TRANSCRIPT = (
    "**Interviewer:** " + "तानाबाना बुनाई की प्रक्रिया बहुत पुरानी है। " * 900
    + "\n**Interviewee:** " + "The warp is set at dawn and the weft follows. " * 900
)

TRANSCRIPT = (
    "## Session 1: the loom\n"
    "**Interviewer:** How long have you worked the loom?\n"
    "---\n"
    "**Interviewee 2**: Forty years. My father taught me. *He never used a machine.*\n"
    "- warp set at dawn\n"
    "- weft after the meal\n"
    "### Dyeing\n"
    "**Interviewer:** And the dye? Ask for the `indigo vat`. 🧵\n"
)

# A control byte, a lone surrogate (a client that cut an emoji in half, or a surrogateescape
# decode), and the Unicode noncharacters. openpyxl's own guard covers only the first of these.
DIRTY_NOTE = "Bell\x07 null\x00 vtab\x0b half-emoji \ud83d noncharacter ￾﷐\U0001FFFE end"


def _sheets():
    return [
        {
            "name": "All records",
            "color": "#5B21B6",
            "columns": ["Note", "Kind", "Count"],
            "rows": [
                # Text openpyxl would otherwise store as a formula and Excel would try to run.
                ["=> comb the weft before the shed opens", "Loom notes", 3],
                ["==== DAY 2 ====", "Process", 4],
                ["#REF!", "Reference", 5],
                [DIRTY_NOTE, "Field notes", 6],
                ["चरखा / spinning wheel 🧵", "Devanagari", 7],
                [None, "", 0],
            ],
            "prose": set(),
        },
        {
            # A name carrying every character Excel refuses in a tab title, and far too long.
            "name": "Media by type / hierarchy [live]: everything we hold",
            "color": "#0F766E",
            "columns": ["Type", "Transcript"],
            "rows": [
                ["audio/mp4", TRANSCRIPT],
                ["audio/mp4", LONG_TRANSCRIPT],
                ["audio/mp4", ""],
                # A rich-text cell that is nothing but an empty run and a run with no real font.
                ["audio/mp4", CellRichText([TextBlock(InlineFont(b=True), " "), ""])],
                ["audio/mp4", CellRichText(["\x00", TextBlock(InlineFont(), "\ud83d")])],
            ],
            "prose": {1},
        },
        {
            # Collides with the sheet above once the illegal characters come out.
            "name": "Media by type  hierarchy live: everything we hold too",
            "color": "#B45309",
            "columns": ["Type", "Notes"],
            "rows": [["image/jpeg", "plain"]],
            "prose": set(),
        },
        {
            # Excel quotes sheet names with apostrophes, so the Overview's jump link has to escape
            # the one inside this name; "History" is a title Excel reserves and refuses outright.
            "name": "Weaver's tools",
            "color": "#BE123C",
            "columns": ["Tool", "Note"],
            "rows": [["Charkha", "fine"]],
            "prose": set(),
        },
        {
            "name": "History",
            "color": "#1D4ED8",
            "columns": ["When", "What"],
            "rows": [["2026", "loom rebuilt"]],
            "prose": set(),
        },
    ]


def _rendered(sheets):
    """What data_browser's ``_rendered`` does: prose columns become Excel rich text."""
    out = []
    for sheet in sheets:
        prose = sheet.get("prose") or set()
        if not prose:
            out.append(sheet)
            continue
        out.append(
            {
                **sheet,
                "rows": [
                    [transcript_cell(v) if i in prose else v for i, v in enumerate(row)]
                    for row in sheet["rows"]
                ],
            }
        )
    return out


@pytest.fixture(scope="module")
def workbook_bytes() -> bytes:
    return build_report_workbook(_rendered(_sheets()), "Repository report")


@pytest.fixture(scope="module")
def archive(workbook_bytes: bytes) -> zipfile.ZipFile:
    return zipfile.ZipFile(BytesIO(workbook_bytes))


@pytest.fixture(scope="module")
def parts(archive: zipfile.ZipFile) -> dict[str, ET.Element]:
    """Every XML part in the archive, parsed. Parsing IS the assertion Excel makes."""
    trees = {}
    for name in archive.namelist():
        if name.endswith((".xml", ".rels")):
            trees[name] = ET.fromstring(archive.read(name))
    return trees


# --------------------------------------------------------------------------------------------
# The package holds together.
# --------------------------------------------------------------------------------------------


def test_every_xml_part_parses(archive: zipfile.ZipFile) -> None:
    for name in archive.namelist():
        if not name.endswith((".xml", ".rels")):
            continue
        try:
            ET.fromstring(archive.read(name))
        except ET.ParseError as exc:  # pragma: no cover - the failure message is the point
            pytest.fail(f"{name} is not well-formed XML, so Excel will offer to recover it: {exc}")


def test_no_illegal_codepoint_reaches_the_xml(archive: zipfile.ZipFile) -> None:
    """Raw or as a numeric reference — either way the part stops being XML."""
    for name in archive.namelist():
        if not name.endswith((".xml", ".rels")):
            continue
        text = archive.read(name).decode("utf-8")
        found = _ILLEGAL.search(text)
        assert found is None, f"{name} carries U+{ord(found.group()):04X} verbatim"
        for ref in _CHAR_REF.finditer(text):
            point = int(ref.group(1))
            assert not _ILLEGAL.match(chr(point)), f"{name} carries &#{point}; (U+{point:04X})"


def test_declared_parts_all_exist(archive: zipfile.ZipFile, parts) -> None:
    names = set(archive.namelist())
    for override in parts["[Content_Types].xml"].findall(f"{CONTENT_TYPES}Override"):
        part = override.get("PartName", "").lstrip("/")
        assert part in names, f"[Content_Types].xml declares {part}, which is not in the archive"


def test_every_relationship_resolves(archive: zipfile.ZipFile, parts) -> None:
    names = set(archive.namelist())
    for name, tree in parts.items():
        if not name.endswith(".rels"):
            continue
        base = posixpath.dirname(posixpath.dirname(name))
        for rel in tree.findall(f"{RELS}Relationship"):
            if rel.get("TargetMode") == "External":
                continue
            target = rel.get("Target", "")
            resolved = (
                target.lstrip("/")
                if target.startswith("/")
                else posixpath.normpath(posixpath.join(base, target))
            )
            assert resolved in names, f"{name} points at {target}, which is missing"


def test_openpyxl_can_read_it_back(workbook_bytes: bytes) -> None:
    """A second parser over the finished bytes, not over our in-memory objects."""
    wb = load_workbook(BytesIO(workbook_bytes), rich_text=True)
    assert wb.sheetnames[0] == "Overview"
    assert len(wb.sheetnames) == len(_sheets()) + 1


# --------------------------------------------------------------------------------------------
# The cells say what Excel can accept.
# --------------------------------------------------------------------------------------------


def _worksheets(parts):
    return {n: t for n, t in parts.items() if n.startswith("xl/worksheets/sheet")}


def test_no_text_cell_was_written_as_a_formula(parts) -> None:
    """A note opening with "=" is text. Written as <f>, Excel evaluates it and then repairs."""
    for name, tree in _worksheets(parts).items():
        for cell in tree.iter(f"{MAIN}c"):
            formula = cell.find(f"{MAIN}f")
            assert formula is None, (
                f"{name} {cell.get('r')} was written as a formula: <f>{formula.text}</f>"
            )


def test_the_equals_sign_the_researcher_typed_survives(workbook_bytes: bytes) -> None:
    """Pinning the type back to text must not mangle the text to dodge the inference."""
    wb = load_workbook(BytesIO(workbook_bytes))
    notes = [row[0] for row in wb["All records"].iter_rows(min_row=2, values_only=True)]
    assert "=> comb the weft before the shed opens" in notes
    assert "==== DAY 2 ====" in notes
    assert "#REF!" in notes


def test_no_cell_exceeds_the_excel_character_ceiling(parts) -> None:
    for name, tree in _worksheets(parts).items():
        for cell in tree.iter(f"{MAIN}c"):
            total = sum(len(t.text or "") for t in cell.iter(f"{MAIN}t"))
            assert total <= MAX_CELL_CHARS, f"{name} {cell.get('r')} holds {total} characters"


def test_no_rich_text_run_is_empty(parts) -> None:
    """openpyxl prunes empty runs on ``+=`` and never on ``append``; <r><t/></r> is not a run."""
    for name, tree in _worksheets(parts).items():
        for cell in tree.iter(f"{MAIN}c"):
            for run in cell.iter(f"{MAIN}r"):
                text = run.find(f"{MAIN}t")
                assert text is not None and text.text, f"{name} {cell.get('r')} has an empty run"


# --------------------------------------------------------------------------------------------
# Sheet titles and the links between them.
# --------------------------------------------------------------------------------------------


def test_sheet_titles_are_legal_and_unique(parts) -> None:
    titles = [s.get("name", "") for s in parts["xl/workbook.xml"].iter(f"{MAIN}sheet")]
    for title in titles:
        assert title, "a sheet was given an empty title"
        assert len(title) <= 31, f"{title!r} is longer than Excel's 31-character limit"
        assert not set(title) & TITLE_BAD, f"{title!r} carries a character Excel forbids"
        assert not title.startswith("'") and not title.endswith("'"), f"{title!r} is quoted"
        assert title.lower() != "history", "'History' is a title Excel reserves"
    lowered = [t.lower() for t in titles]
    assert len(set(lowered)) == len(lowered), f"duplicate sheet titles: {titles}"


def test_overview_links_point_at_sheets_that_exist(parts) -> None:
    """A name like "Weaver's tools" has to have its apostrophe doubled or the link resolves to
    nothing, because Excel wraps a sheet reference in apostrophes of its own."""
    titles = {s.get("name", "") for s in parts["xl/workbook.xml"].iter(f"{MAIN}sheet")}
    links = list(parts["xl/worksheets/sheet1.xml"].iter(f"{MAIN}hyperlink"))
    assert links, "the Overview lost its jump links"
    for link in links:
        assert link.get("ref"), "a hyperlink was written without a cell reference"
        location = link.get("location", "")
        quoted, _, remainder = location.partition("!")
        assert remainder == "A1", f"{location!r} does not target a cell"
        assert quoted.startswith("'") and quoted.endswith("'"), f"{location!r} is unquoted"
        inner = quoted[1:-1]
        # Every apostrophe inside the quotes must be doubled. Checking only that the name maps
        # back to a sheet is not enough: strip the outer quotes off "'Weaver's tools'" and you get
        # the right name by accident, while Excel has already stopped reading at the third quote.
        assert re.fullmatch(r"(?:[^']|'')*", inner), f"{location!r} has a bare apostrophe"
        assert inner.replace("''", "'") in titles, f"{location!r} names no sheet"


# --------------------------------------------------------------------------------------------
# The transcript reads as dialogue rather than as Markdown source.
# --------------------------------------------------------------------------------------------


def test_transcript_renders_as_formatting_not_as_markdown() -> None:
    rich = transcript_cell(TRANSCRIPT)
    plain = str(rich)
    assert "**" not in plain, "the asterisks are still in the cell"
    assert not plain.lstrip().startswith("#"), "the heading is still written with hashes"
    assert "## " not in plain and "### " not in plain
    assert "`" not in plain, "a code span kept its backticks"
    assert "•  warp set at dawn" in plain, "a list item did not become a bullet"
    assert "Session 1: the loom" in plain and "Dyeing" in plain
    assert "\n" in plain, "every turn is on one line, so the cell reads as a blob"

    bold = {b.text for b in rich if isinstance(b, TextBlock) and b.font.b}
    assert "Interviewer:" in bold, "the speaker label is not bold"
    assert "Session 1: the loom" in bold and "Dyeing" in bold, "headings are not bold"
    italic = {b.text for b in rich if isinstance(b, TextBlock) and b.font.i}
    assert "He never used a machine." in italic


def test_underscores_are_left_alone() -> None:
    """Transcripts carry object keys and file names; italicising them would eat the data."""
    assert "media_file_01_raw.wav" in str(transcript_cell("Saved as media_file_01_raw.wav"))


def test_prose_cells_are_wrapped_and_given_room(workbook_bytes: bytes) -> None:
    wb = load_workbook(BytesIO(workbook_bytes))
    # By position, not by name: the sheet's title is whatever survived _safe_title, and that is
    # the point of the sheet — Overview first, then the sheets in the order they were handed over.
    ws = wb.worksheets[2]
    cell = ws["B2"]
    assert cell.alignment.wrap_text, "a transcript without wrap_text shows as one crushed line"
    assert ws.column_dimensions["B"].width >= 60, "the transcript column is too narrow to read"
    assert (ws.row_dimensions[2].height or 0) > 15, "the transcript row was left one line tall"
    assert (ws.row_dimensions[2].height or 0) <= 409, "Excel rejects a row taller than 409"
