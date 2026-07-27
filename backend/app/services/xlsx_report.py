"""Styled .xlsx workbook builder for the /data/report relational report.

The data_browser route assembles the report as a list of plain-dict *sheets* (one per record
type) and hands them here to be rendered into an Excel workbook:

    sheet = {"name": str, "color": "#RRGGBB", "columns": [label, ...], "rows": [[cell, ...], ...]}

``build_report_workbook`` turns those into a workbook with:

- a leading **Overview** sheet listing every data sheet, its row count, and an internal
  hyperlink that jumps to it,
- per-sheet styling: a bold white header row on the sheet's brand colour, a matching tab colour,
  thin header borders, a frozen header row (``freeze_panes = "A2"``) and column widths sized to
  content (capped so a transcript column can't blow the layout out),
- prose cells: a value that is rich text (``CellRichText``, as ``transcript_format`` builds) or
  simply carries newlines is laid out to be read rather than glanced at — wrapped, top-aligned,
  in a wide column, on a row grown to fit it but never past ``_MAX_WRAP_LINES``.

Every value reaches its cell through ``_put`` and nothing else. That single door is what keeps
Excel from opening the download with "We found a problem with some content" and offering to
recover it, which it did for three separate reasons, all of them reachable from ordinary field
data:

1. **Codepoints XML cannot carry.** A worksheet part is XML and XML 1.0 admits only #x9, #xA,
   #xD and the printable ranges. openpyxl guards the C0 block and nothing else, so a lone
   surrogate — a phone that cut an emoji in half, a string decoded with ``surrogateescape`` —
   reached the writer, which serialised it as ``&#55357;``: a numeric reference to a codepoint
   no XML parser will accept. The part stopped being well-formed at all.
2. **Field notes read as formulas.** openpyxl infers a cell's type from its text: anything
   opening with ``=`` is stored as a formula, anything spelling ``#REF!`` as an error. A note
   beginning "=> comb the weft first" was written as ``<f>&gt; comb the weft first</f>``, which
   Excel dutifully tried to evaluate and then offered to repair.
3. **Sheet references that don't resolve.** The Overview's jump links quote the target sheet's
   name in single quotes, so a workshop called "Weaver's tools" ended the quoted span early.

The same door also clips to Excel's 32767-character cell ceiling and keeps worksheet titles
legal and unique. Adding a value to this workbook without going through ``_put`` is how all
three of the above got in; ``backend/tests/test_xlsx_report.py`` fails if any of them come back.

openpyxl is synchronous; the caller runs this inside ``asyncio.to_thread``. The single public
entry point returns the finished workbook as ``bytes``.
"""

import re
from io import BytesIO
from math import ceil
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Excel's hard per-cell character ceiling; longer values corrupt the file.
_MAX_CELL_CHARS = 32767
# Column width is measured in characters.
_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 10
# A transcript is read as a block of dialogue, so its column gets more room than any data column.
_PROSE_COL_WIDTH = 80
# Excel worksheet-title rules: <= 31 chars, none of these characters.
_TITLE_MAX = 31
_TITLE_BAD = set(r"[]:*?/\\")
# Excel keeps "History" for itself (the shared-workbook change log) and refuses a sheet by that
# name; a craft or workshop legitimately called that would otherwise fail the save.
_TITLE_RESERVED = {"history"}

_OVERVIEW_NAME = "Overview"
_OVERVIEW_COLOR = "5B21B6"  # purple, matching the Workshops brand tone

_WHITE = "FFFFFFFF"
_HYPERLINK_BLUE = "FF1D4ED8"

_THIN = Side(style="thin", color="FFBFBFBF")
_HEADER_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
_CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)
_WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
# Left to auto-fit, a wrapped transcript grows a row hundreds of lines tall and the sheet stops
# being scrollable. Size wrapped rows to their own text instead, up to a readable window; past
# that the reader drags the row open.
_LINE_HEIGHT = 15
_MAX_WRAP_LINES = 16
# Excel's own ceiling on a row; anything taller is rejected outright.
_MAX_ROW_HEIGHT = 409.0

# The noncharacter pair at the top of every supplementary plane. Valid Python string content,
# valid UTF-8, and rejected by Excel's XML parser — so they are stripped here rather than left to
# surface as a repair prompt on someone's laptop in the field.
_NONCHARACTERS = "".join(
    chr(plane * 0x10000 + offset) for plane in range(1, 17) for offset in (0xFFFE, 0xFFFF)
)
# Everything XML 1.0 forbids inside element content. This is deliberately wider than openpyxl's
# own ILLEGAL_CHARACTERS_RE, which stops at the C0 block: the characters that actually broke the
# download were a lone surrogate and a noncharacter, neither of which openpyxl looks at. Tab,
# newline and carriage return are kept — a transcript is nothing but newlines.
_XML_UNSAFE_RE = re.compile(
    "["
    "\x00-\x08\x0b\x0c\x0e-\x1f"  # C0 controls, less \t \n \r
    "\ud800-\udfff"  # lone surrogates, e.g. an emoji truncated mid-pair by a client
    "\ufdd0-\ufdef\ufffe\uffff"  # noncharacters in the BMP
    + _NONCHARACTERS
    + "]"
)

# What a run falls back to when its font reference is missing or is not an InlineFont: plain
# text is a far better outcome than a 500 on the download.
_PLAIN_FONT = InlineFont()


def _rgb(color: str | None) -> str:
    """'#5B21B6' / '5b21b6' -> 'FF5B21B6' (openpyxl wants opaque 8-hex ARGB)."""
    hexpart = (color or "").lstrip("#").strip().upper()
    if len(hexpart) == 6:
        return f"FF{hexpart}"
    if len(hexpart) == 8:
        return hexpart
    return f"FF{_OVERVIEW_COLOR}"


def _tab_color(color: str | None) -> str:
    # Full opaque ARGB: a bare 6-hex value gets a 00 (transparent) alpha and the tab shows blank.
    return _rgb(color)


def _sanitise(text: str) -> str:
    """Text an Excel cell can hold: no codepoint XML rejects, no more than the cell ceiling.

    Stripping rather than rejecting, for the same reason ``records.contains`` strips NUL rather
    than raising: a researcher who pasted a name out of a PDF and picked up a byte they cannot
    see wants their report, not a failed download.
    """
    text = _XML_UNSAFE_RE.sub("", text)
    if len(text) > _MAX_CELL_CHARS:
        text = text[: _MAX_CELL_CHARS - 1] + "…"
    return text


def _clean(value: Any) -> Any:
    """Make a value safe for an Excel cell: strip illegal codepoints, clip over-long text."""
    if value is None:
        return None
    if isinstance(value, CellRichText):
        # A rich-text value whose every run sanitised away becomes a plain blank: an empty
        # CellRichText still writes an <is/> with no text in it, which is not what a blank cell is.
        return _clean_runs(value) or ""
    if isinstance(value, (int, float)):
        return value
    return _sanitise(value if isinstance(value, str) else str(value))


def _clean_runs(value: CellRichText) -> CellRichText:
    """``_clean`` for rich text — same guarantees, but the cell limit spans all of its runs.

    Rich text is how the longest cells in the workbook arrive, so the clip has to be honoured
    across the runs rather than per run; a run cut short ends the cell. Runs that sanitise down
    to nothing are dropped: openpyxl only prunes empty runs on ``+=``, never on ``append``, so an
    empty one would otherwise be written out as ``<r><t/></r>``. A cell whose every run vanishes
    comes back as a plain empty string rather than an empty rich-text object.
    """
    cleaned = CellRichText()
    budget = _MAX_CELL_CHARS
    for run in value:
        block = isinstance(run, TextBlock)
        text = _XML_UNSAFE_RE.sub("", run.text if block else str(run))
        if not text:
            continue
        if len(text) >= budget:
            text = text[: budget - 1] + "…"
            budget = 0
        else:
            budget -= len(text)
        if block:
            font = run.font if isinstance(run.font, InlineFont) else _PLAIN_FONT
            cleaned.append(TextBlock(font, text))
        else:
            cleaned.append(text)
        if not budget:
            break
    return cleaned


def _put(ws: Worksheet, row: int, column: int, value: Any) -> Cell:
    """Write one cell. Every value in this workbook goes through here and nowhere else.

    Beyond sanitising, this pins a text cell back to being text. openpyxl reads the *meaning* of
    a string as it binds it — a leading "=" makes the cell a formula, "#REF!" makes it an error —
    and Excel then tries to evaluate the researcher's field note and offers to repair the file
    when it cannot. This report never emits a formula, so there is nothing to lose by saying so
    outright, and the note keeps the "=" the researcher typed instead of being mangled to dodge
    the inference.
    """
    cell = ws.cell(row=row, column=column, value=_clean(value))
    if isinstance(cell.value, str) and cell.data_type != "s":
        cell.data_type = "s"
    return cell


def _wraps(value: Any) -> bool:
    """Whether a cell holds prose: rich text, or text the author already broke into lines."""
    if isinstance(value, CellRichText):
        return True
    return isinstance(value, str) and "\n" in value


def _wrap_height(text: str) -> float:
    """How tall a wrapped cell needs to sit: its own lines, each re-broken to the column width."""
    lines = sum(max(1, ceil(len(line) / _PROSE_COL_WIDTH)) for line in text.split("\n"))
    return min(min(lines, _MAX_WRAP_LINES) * _LINE_HEIGHT, _MAX_ROW_HEIGHT)


def _safe_title(name: str, used: set[str]) -> str:
    """A legal, unique worksheet title for ``name``.

    Excel's rules, all of which a workshop or craft name typed by a researcher can break: at most
    31 characters, none of ``[]:*?/\\``, no leading or trailing apostrophe (it quotes sheet names
    with those), not "History", and unique case-insensitively across the workbook.
    """
    base = "".join(ch for ch in _sanitise(name or "").strip() if ch not in _TITLE_BAD)
    base = base[:_TITLE_MAX].strip().strip("'").strip()
    if not base or base.lower() in _TITLE_RESERVED:
        base = f"{base} sheet".strip() if base else "Sheet"
    candidate = base
    n = 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = base[: _TITLE_MAX - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _sheet_ref(title: str) -> str:
    """``Weaver's tools`` -> ``'Weaver''s tools'!A1``, the jump target for an Overview link.

    Excel quotes a sheet name in a reference with single quotes, so a name carrying one of its
    own closes the quoted span early and the link resolves to nothing. Doubling is Excel's escape.
    """
    return "'{}'!A1".format(title.replace("'", "''"))


def _autosize(ws: Worksheet, columns: list[str], rows: list[list[Any]], wrapped: set[int]) -> None:
    for idx, header in enumerate(columns):
        if idx in wrapped:
            # Prose is read down the column, not across it — always give it the full allowance.
            ws.column_dimensions[get_column_letter(idx + 1)].width = _PROSE_COL_WIDTH
            continue
        longest = len(str(header))
        for row in rows:
            if idx < len(row) and row[idx] is not None:
                # Widen to the longest single line only — newlines shouldn't stretch the column.
                cell = str(row[idx])
                line = max((len(part) for part in cell.splitlines()), default=len(cell))
                longest = max(longest, line)
        width = max(_MIN_COL_WIDTH, min(longest + 2, _MAX_COL_WIDTH))
        ws.column_dimensions[get_column_letter(idx + 1)].width = width


def _write_sheet(ws: Worksheet, sheet: dict[str, Any]) -> None:
    columns: list[str] = sheet.get("columns") or []
    rows: list[list[Any]] = sheet.get("rows") or []
    fill = PatternFill(fill_type="solid", start_color=_rgb(sheet.get("color")))
    ws.sheet_properties.tabColor = _tab_color(sheet.get("color"))

    for idx, label in enumerate(columns, start=1):
        cell = _put(ws, 1, idx, label)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = fill
        cell.border = _HEADER_BORDER
        cell.alignment = _HEADER_ALIGN

    wrapped: set[int] = set()
    for r, row in enumerate(rows, start=2):
        height = 0.0
        for idx, value in enumerate(row):
            cell = _put(ws, r, idx + 1, value)
            if _wraps(cell.value):
                cell.alignment = _WRAP_ALIGN
                wrapped.add(idx)
                height = max(height, _wrap_height(str(cell.value)))
            else:
                cell.alignment = _CELL_ALIGN
        if height:
            ws.row_dimensions[r].height = height

    if columns:
        ws.freeze_panes = "A2"
    _autosize(ws, columns, rows, wrapped)


def _write_overview(ws: Worksheet, title: str, sheets: list[dict[str, Any]]) -> None:
    ws.sheet_properties.tabColor = _rgb(_OVERVIEW_COLOR)

    heading = _put(ws, 1, 1, title or "Report")
    heading.font = Font(bold=True, size=14, color=f"FF{_OVERVIEW_COLOR}")

    headers = ["Sheet", "Rows"]
    fill = PatternFill(fill_type="solid", start_color=_rgb(_OVERVIEW_COLOR))
    for idx, label in enumerate(headers, start=1):
        cell = _put(ws, 3, idx, label)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = fill
        cell.border = _HEADER_BORDER
        cell.alignment = _HEADER_ALIGN

    for offset, sheet in enumerate(sheets):
        r = 4 + offset
        safe = sheet["_title"]
        link = _put(ws, r, 1, sheet.get("name") or safe)
        link.hyperlink = Hyperlink(ref="", location=_sheet_ref(safe), display=safe)
        link.font = Font(color=_HYPERLINK_BLUE, underline="single")
        _put(ws, r, 2, len(sheet.get("rows") or []))

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.freeze_panes = "A4"


def build_report_workbook(sheets: list[dict[str, Any]], title: str) -> bytes:
    """Render the report ``sheets`` into a styled .xlsx workbook and return its bytes.

    ``sheets`` is the list produced by data_browser's ``_report_sheets`` — each a dict of
    ``{name, color (#hex), columns, rows}``. A leading Overview sheet is prepended.
    """
    wb = Workbook()

    used_titles: set[str] = {_OVERVIEW_NAME.lower()}
    for sheet in sheets:
        sheet["_title"] = _safe_title(sheet.get("name") or "Sheet", used_titles)

    overview = wb.active
    overview.title = _OVERVIEW_NAME
    _write_overview(overview, title, sheets)

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet["_title"])
        _write_sheet(ws, sheet)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
